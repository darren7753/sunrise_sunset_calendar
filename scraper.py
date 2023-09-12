# import chromedriver_autoinstaller
import time
import os
import pandas as pd
import numpy as np

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from openpyxl import load_workbook
from collections import defaultdict
from bs4 import BeautifulSoup

data_folder = "Data"
if not os.path.exists(data_folder):
    os.makedirs(data_folder)

sheet_names = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

def select_checkbox_or_radio_by_id(element_id):
    element = driver.find_element(By.ID, element_id)
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    
    if not element.is_selected():
        element.click()

def get_moon_phase(group):
    if "Full Moon" in group["Data"].values:
        group["Moon Phases"] = 1
    elif "First Qtr" in group["Data"].values or "Last Qtr" in group["Data"].values:
        group["Moon Phases"] = 2
    elif "New Moon" in group["Data"].values:
        group["Moon Phases"] = 3
    else:
        group["Moon Phases"] = 0
    return group

# chromedriver_autoinstaller.install()

url = "http://sunrisesunset.com/USA/"

extension_path_1 = "extension_3_4_0_0.crx"
extension_path_2 = "extension_5_10_0_0.crx"

options = Options()
options.add_extension(extension_path_1)
options.add_extension(extension_path_2)
options.add_argument("--headless")
options.add_argument("window-size=1920x1080")

driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.get(url)

time.sleep(3)

main_tab = None
for tab in driver.window_handles:
    driver.switch_to.window(tab)
    if driver.current_url == url:
        main_tab = tab
        break

if main_tab is None:
    raise Exception("Main tab not found!")

for tab in driver.window_handles:
    if tab != main_tab:
        driver.switch_to.window(tab)
        driver.close()

driver.switch_to.window(main_tab)

links = driver.find_elements(By.XPATH, "//a[starts-with(@href, '/USA/') and not(contains(@href, 'print'))]")

start_navigating = False

for i in range(len(links)):
    links = driver.find_elements(By.XPATH, "//a[starts-with(@href, '/USA/') and not(contains(@href, 'print'))]")
    link = links[i]
    
    if "Nebraska" in link.get_attribute("href"):
        start_navigating = True

    if start_navigating:
        href = link.get_attribute("href")
        print("Visiting link:", href)
        state_name = link.text

        year_month_mapping = {
            "2022": list(range(1, 13)),
            "2023": list(range(1, 5))
        }

        for year, _ in year_month_mapping.items():
            with pd.ExcelWriter(f"{data_folder}/{state_name}_{year}.xlsx", engine="openpyxl") as writer:
                for sheet in sheet_names:
                    df_empty = pd.DataFrame()
                    df_empty.to_excel(writer, sheet_name=sheet, index=False)

        driver.execute_script("arguments[0].scrollIntoView();", link)
        link.click()

        try:
            dropdown_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "comb_city_info")))
            dropdown = Select(dropdown_element)

            for n, option in enumerate(dropdown.options):
                if option.text == "Choose a location":
                    continue

                print(f"Selecting location: {option.text}")
                city_name = option.text
                dropdown.select_by_visible_text(option.text)
                
                checkbox_and_radio_ids = [
                    "want_twi_civ", 
                    "want_twi_naut", 
                    "want_twi_astro", 
                    "want_info", 
                    "want_mphase", 
                    "want_mrms", 
                    "want_solar_noon", 
                    "want_eqx_sol", 
                    "time_24hr", 
                    "wsos_yes"
                ]

                year_input = driver.find_element(By.ID, "year")
                month_dropdown = Select(driver.find_element(By.NAME, "month"))
                
                for year, months in year_month_mapping.items():
                    year_input.clear()
                    year_input.send_keys(year)
                    
                    for month_value in months:
                        month_dropdown.select_by_value(str(month_value))
                        month_name = month_dropdown.first_selected_option.text
                        
                        for element_id in checkbox_and_radio_ids:
                            select_checkbox_or_radio_by_id(element_id)

                        form_element = driver.find_element(By.TAG_NAME, "form")

                        driver.execute_script("arguments[0].setAttribute('target', '_blank');", form_element)

                        submit_button = driver.find_element(By.XPATH, '//input[@type="submit" and @value="Make Calendar"]')
                        submit_button.click()

                        driver.switch_to.window(driver.window_handles[-1])

                        soup = BeautifulSoup(driver.page_source, 'html.parser')
                        table = soup.select_one("table[style='width:96%; margin-left:4px; margin-bottom:10px; border-collapse:collapse; border-spacing:0; border:2px solid black; ']")
                        data = []
                        for row in table.select("tr")[1:]:
                            cells = row.select("td")
                            for cell in cells:
                                day_data = {}
                                day_number_elements = cell.select("span.daynum")
                                if day_number_elements:
                                    day_data["Day"] = day_number_elements[0].text
                                    events = cell.get_text('\n').split('\n')[1:]
                                    event_dict = defaultdict(list)
                                    for event in events:
                                        if event and ": " in event:
                                            event_parts = event.split(": ")
                                            event_dict[event_parts[0]].append(event_parts[1])
                                    for k, v in event_dict.items():
                                        day_data[k] = ', '.join(v)
                                    day_data["City"] = city_name
                                    day_data["State"] = state_name
                                    data.append(day_data)

                        df = pd.DataFrame(data)

                        for col in df.columns:
                            if df[col].dtype == "object" and df[col].str.contains(",").any():
                                splits = df[col].str.split(",", expand=True)
                                
                                for i in range(splits.shape[1]):
                                    df[f"{col}_{i+1}"] = splits[i]
                                    
                                df.drop(col, axis=1, inplace=True)

                        df_melted = df.melt(id_vars=["Day", "State", "City"], 
                                            value_vars=list(df.drop(["Day", "State", "City"], axis=1).columns),
                                            var_name="Data",
                                            value_name="Time")
                        df_melted = df_melted.replace("none", np.nan)
                        df_melted.dropna(inplace=True)
                        df_melted = df_melted.groupby("Day").apply(get_moon_phase).reset_index(drop=True)
                        df_melted["Day"] = pd.to_datetime(str(month_value) + "-" + df_melted["Day"].astype(str) + "-" + str(year)).dt.date
                        df_melted["Time"] = df_melted["Time"].str.strip().replace("24:00", "00:00")
                        df_melted["Time"] = pd.to_datetime(df_melted["Time"], format="%H:%M").dt.time
                        df_melted.rename(columns={"Day": "Date"}, inplace=True)
                        df_melted = df_melted[["Date", "State", "City", "Moon Phases", "Data", "Time"]]
                        df_melted = df_melted.sort_values(["Date", "State", "City", "Time"])
                        df_melted["Data"] = df_melted["Data"].str.split("_").str[0]
                        df_melted.reset_index(drop=True, inplace=True)

                        print(df_melted.head(10))

                        wb = load_workbook(f"{data_folder}/{state_name}_{year}.xlsx")
                        ws = wb[month_name]

                        if n == 1:
                            if ws.max_row == 1: 
                                ws.append(df_melted.columns.tolist())

                        for index, row in df_melted.iterrows():
                            ws.append(row.values.tolist())

                        wb.save(f"{data_folder}/{state_name}_{year}.xlsx")

                        print(f"The data has been stored in the {month_name} sheet of {data_folder}/{state_name}_{year}.xlsx")

                        driver.close()
                        driver.switch_to.window(main_tab)

        except NoSuchElementException:
            print(f"Dropdown not found on {href}")
        except Exception as e:
            print(f"Error encountered: {e}")

        driver.back()

driver.quit()